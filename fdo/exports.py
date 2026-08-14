"""Executive deliverables: a PDF briefing (matplotlib PdfPages) and an Excel
workbook (openpyxl). Every number in both files comes from one
``run_pipeline`` result, so the deliverables can never drift from the code.

The PDF states on its cover that the data is synthetic and the costs are
assumptions; the Excel workbook carries a dedicated Assumptions sheet.
"""

from __future__ import annotations

import os

import matplotlib

matplotlib.use("Agg")

import matplotlib.pyplot as plt
import numpy as np
from matplotlib.backends.backend_pdf import PdfPages
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# Palette lives in fdo/palette.py (fraud review floor: paper neutrals, alert
# amber, confirmed-fraud red, cleared green; the validator run is recorded
# there). Cleared green and the stone fill are sub-3:1 on the light print
# surface, so every mark drawn in them ships with a visible direct label.
from fdo.palette import (
    ALERT_AMBER,
    BASELINE,
    CASE_GRAPHITE,
    CLEARED_GREEN,
    GRID,
    INK,
    INK_2,
    MUTED,
    PAPER_STONE,
    SURFACE,
)
from fdo.pipeline import headline

MIN_BYTES = 10_240


def _esc(s: str) -> str:
    """Escape dollar signs so matplotlib never parses $...$ spans as math
    (log-axis tick labels legitimately use mathtext, so the global rcParam
    switch is not an option)."""
    return s.replace("$", "\\$")


def _style_axes(ax: plt.Axes) -> None:
    ax.set_facecolor(SURFACE)
    for side in ("top", "right"):
        ax.spines[side].set_visible(False)
    for side in ("left", "bottom"):
        ax.spines[side].set_color(BASELINE)
    ax.grid(True, color=GRID, linewidth=0.8)
    ax.set_axisbelow(True)
    ax.tick_params(colors=MUTED, labelsize=9)
    ax.xaxis.label.set_color(INK_2)
    ax.yaxis.label.set_color(INK_2)
    ax.title.set_color(INK)


def _cover_page(pdf: PdfPages, results: dict) -> None:
    fig = plt.figure(figsize=(11, 8.5))
    fig.patch.set_facecolor(SURFACE)
    fig.text(0.07, 0.90, "Fraud Detection Ops", fontsize=26, color=INK, weight="bold")
    fig.text(
        0.07, 0.855,
        "Class-imbalanced fraud scoring wired to the analyst review queue",
        fontsize=13, color=INK_2,
    )
    fig.text(
        0.07, 0.775,
        "DISCLAIMER - read first\n"
        "All transactions are SYNTHETIC, produced by a seeded generator (fdo/data.py). No real\n"
        "customer or merchant data is used. All dollar costs in the threshold and queue analyses\n"
        "are labelled ASSUMPTIONS, not measurements. Numbers below are measured on a held-out\n"
        "time-based test window the model never trained or calibrated on.",
        fontsize=10, color=INK_2, va="top",
        bbox={"boxstyle": "round,pad=0.6", "facecolor": "white",
              "edgecolor": BASELINE, "linewidth": 1.0},
    )
    fig.text(0.07, 0.60, "Headline results", fontsize=14, color=INK, weight="bold")
    y = 0.565
    for line in headline(results):
        fig.text(0.07, y, _esc("-  " + line), fontsize=10.5, color=INK, va="top")
        y -= 0.052
    fig.text(
        0.07, 0.10,
        "Everything implemented from scratch in NumPy (no scikit-learn / XGBoost); "
        "queue allocation via SciPy HiGHS.\n"
        "Dimitres Kisimov - 2026 - all rights reserved (published for portfolio review).",
        fontsize=9, color=MUTED,
    )
    pdf.savefig(fig)
    plt.close(fig)


def _pr_page(pdf: PdfPages, results: dict) -> None:
    m = results["metrics"]
    prec, rec = results["plots"]["pr_model_test"]
    fig, ax = plt.subplots(figsize=(11, 8.5))
    fig.patch.set_facecolor(SURFACE)
    _style_axes(ax)
    ax.plot(rec, prec, color=ALERT_AMBER, linewidth=2.0,
            label=f"Model ({m['primary_model']}) - PR-AUC {m['test_primary']['pr_auc']:.3f}")
    prev = m["baseline_random"]["theoretical_pr_auc"]
    ax.axhline(prev, color=CASE_GRAPHITE, linewidth=2.0, linestyle="--",
               label=f"Prevalence-random floor - PR-AUC {m['baseline_random']['pr_auc']:.3f}")
    bp = m["baseline_p99"]
    ax.plot([bp["flag_recall"]], [bp["flag_precision"]], marker="o", markersize=9,
            color=CASE_GRAPHITE, markeredgecolor=INK, markeredgewidth=0.8, linestyle="none",
            label=f"Rule: amount > p99 - PR-AUC {bp['pr_auc']:.3f}")
    ax.annotate("amount > p99 rule", (bp["flag_recall"], bp["flag_precision"]),
                textcoords="offset points", xytext=(10, 6), fontsize=9, color=INK_2)
    ax.annotate(f"random floor = prevalence ({prev:.3f})", (0.62, prev),
                textcoords="offset points", xytext=(0, 6), fontsize=9, color=INK_2)
    ax.set_xlabel("Recall (share of fraud caught)")
    ax.set_ylabel("Precision (share of flags that are fraud)")
    ax.set_title(
        "Precision-recall on the held-out test window - why accuracy is the wrong metric\n"
        f"(at {m['test_primary']['prevalence']:.2%} prevalence, 'always legitimate' is "
        f"{1 - m['test_primary']['prevalence']:.1%} accurate and catches nothing; "
        f"oracle ceiling on this synthetic data: PR-AUC {m['oracle_test_pr_auc']:.3f})",
        fontsize=11,
    )
    leg = ax.legend(loc="upper right", frameon=True, fontsize=10)
    leg.get_frame().set_edgecolor(BASELINE)
    ax.set_xlim(0, 1)
    ax.set_ylim(0, 1)
    pdf.savefig(fig)
    plt.close(fig)


def _reliability_page(pdf: PdfPages, results: dict) -> None:
    cal = results["metrics"]["calibration"]
    raw = results["plots"]["reliability_val_raw"]
    calc = results["plots"]["reliability_val_cal"]
    fig, ax = plt.subplots(figsize=(11, 8.5))
    fig.patch.set_facecolor(SURFACE)
    _style_axes(ax)
    ax.plot([0, 1], [0, 1], color=BASELINE, linewidth=1.2, linestyle=":", zorder=1)
    ax.annotate("perfect calibration", (0.66, 0.66), rotation=38, fontsize=9,
                color=MUTED, va="bottom")
    ax.plot(raw["confidence"], raw["accuracy"], color=ALERT_AMBER, linewidth=2.0,
            marker="o", markersize=8, markeredgecolor=SURFACE, markeredgewidth=1.5,
            label=f"Before calibration - ECE {cal['val_ece_raw']:.3f}")
    ax.plot(calc["confidence"], calc["accuracy"], color=CLEARED_GREEN, linewidth=2.0,
            marker="o", markersize=8, markeredgecolor=SURFACE, markeredgewidth=1.5,
            label=f"After Platt scaling - ECE {cal['val_ece_cal']:.3f}")
    ax.set_xlabel("Mean predicted fraud probability (equal-count bin)")
    ax.set_ylabel("Observed fraud rate (bin)")
    ax.set_title(
        "Reliability diagram, validation - class weighting distorts probabilities on purpose\n"
        f"Platt scaling repairs them: Brier {cal['val_brier_raw']:.4f} -> "
        f"{cal['val_brier_cal']:.4f} (val); ECE {cal['test_ece_raw']:.3f} -> "
        f"{cal['test_ece_cal']:.3f} (test)\n"
        "Downstream cost thresholds and queue expected values consume these probabilities",
        fontsize=11,
    )
    leg = ax.legend(loc="upper left", frameon=True, fontsize=10)
    leg.get_frame().set_edgecolor(BASELINE)
    ax.set_xlim(0, 1)
    ax.set_ylim(0, 1)
    pdf.savefig(fig)
    plt.close(fig)


def _threshold_page(pdf: PdfPages, results: dict) -> None:
    t = results["thresholds"]
    sweep = t["sweep_val"]
    fig, ax = plt.subplots(figsize=(11, 8.5))
    fig.patch.set_facecolor(SURFACE)
    _style_axes(ax)
    s = sweep[sweep["threshold"] > 0].sort_values("threshold")
    ax.plot(s["threshold"], s["total_cost"], color=ALERT_AMBER, linewidth=2.0)
    ax.set_xscale("log")
    t_star = t["threshold_star"]
    cost_star = t["val_at_star"]["total_cost"]
    naive = s.iloc[(s["threshold"] - 0.5).abs().argmin()]
    ax.axvline(t_star, color=INK, linewidth=1.2, linestyle="--")
    ax.annotate(
        _esc(f"chosen t* = {t_star:.3f}\nval cost ${cost_star:,.0f}"),
        (t_star, cost_star), textcoords="offset points", xytext=(-150, 45),
        fontsize=10, color=INK,
        arrowprops={"arrowstyle": "-", "color": BASELINE},
    )
    ax.axvline(0.5, color=MUTED, linewidth=1.2, linestyle="--")
    ax.annotate(
        _esc(f"naive default 0.5\nval cost ${naive['total_cost']:,.0f}"),
        (0.5, naive["total_cost"]), textcoords="offset points", xytext=(-30, 60),
        fontsize=10, color=INK_2,
        arrowprops={"arrowstyle": "-", "color": BASELINE},
    )
    ax.set_xlabel("Alert threshold on calibrated fraud probability (log scale)")
    ax.set_ylabel(_esc("Empirical cost on validation ($, under stated assumptions)"))
    ax.set_title(
        _esc(
            "Expected-cost threshold sweep (validation) - threshold chosen here, judged on test\n"
            f"Test window: ${t['test_at_star']['total_cost']:,.0f} at t* vs "
            f"${t['test_at_naive_05']['total_cost']:,.0f} at naive 0.5 "
            f"({t['test_cost_reduction_vs_05']:.1%} lower) vs "
            f"${t['test_review_none']['total_cost']:,.0f} reviewing nothing.\n"
            f"ASSUMPTIONS: ${t['costs'].review_cost:.0f} per review; missed fraud costs its amount."
        ),
        fontsize=11,
    )
    pdf.savefig(fig)
    plt.close(fig)


def _queue_page(pdf: PdfPages, results: dict) -> None:
    q = results["queue"]
    comp = q["comparison"]
    cov = q["coverage"]
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(11, 8.5))
    fig.patch.set_facecolor(SURFACE)
    for ax in (ax1, ax2):
        _style_axes(ax)

    labels = ["Constrained\noptimizer", "Top-K by\nprobability", "Top-K by\nexp. value", "Random"]
    vals = comp["expected_recovered"].to_numpy()
    colors = [ALERT_AMBER, PAPER_STONE, PAPER_STONE, PAPER_STONE]
    bars = ax1.bar(labels, vals, width=0.62, color=colors, edgecolor=SURFACE, linewidth=1.5)
    for rect, v in zip(bars, vals, strict=True):
        ax1.annotate(_esc(f"${v:,.0f}"), (rect.get_x() + rect.get_width() / 2, v),
                     ha="center", va="bottom", fontsize=10, color=INK)
    ax1.set_ylabel(_esc("Expected recovered value, test shift ($)"))
    ax1.set_title(
        f"Who gets reviewed: {q['config'].capacity} reviews of capacity\n"
        "(one measure across strategies; optimizer emphasized)",
        fontsize=11,
    )
    ax1.grid(axis="x", visible=False)
    ax1.tick_params(axis="x", labelsize=9, colors=INK_2)

    x = np.arange(len(cov))
    w = 0.38
    ax2.bar(x - w / 2, cov["milp_reviews"], width=w, color=ALERT_AMBER,
            edgecolor=SURFACE, linewidth=1.5, label="Constrained optimizer")
    ax2.bar(x + w / 2, cov["topk_prob_reviews"], width=w, color=CLEARED_GREEN,
            edgecolor=SURFACE, linewidth=1.5, label="Top-K by probability")
    ax2.plot(x, cov["floor"], linestyle="none", marker="_", markersize=14,
             color=INK, label="Coverage floor (assumed)")
    ax2.set_xticks(x)
    ax2.set_xticklabels(cov["segment"], rotation=35, ha="right", fontsize=8.5, color=INK_2)
    ax2.set_ylabel("Reviews allocated to segment")
    ax2.set_title(
        "Segment coverage - top-K goes blind where\nthe optimizer keeps a floor",
        fontsize=11,
    )
    leg = ax2.legend(loc="upper left", frameon=True, fontsize=9)
    leg.get_frame().set_edgecolor(BASELINE)
    ax2.grid(axis="x", visible=False)

    note = (
        "HONESTY NOTE: without the per-segment coverage floors the LP reduces exactly to\n"
        "'sort by p x amount, take top K' (verified numerically each run: "
        f"{'holds' if q['lp_equals_topk_ev'] else 'DID NOT HOLD'}). "
        "The solver earns its keep once coverage constraints are on."
    )
    fig.text(0.07, 0.03, note, fontsize=9, color=INK_2)
    fig.tight_layout(rect=(0, 0.06, 1, 1))
    pdf.savefig(fig)
    plt.close(fig)


def _fairness_page(pdf: PdfPages, results: dict) -> None:
    from fdo.fairness import queue_fairness

    f = queue_fairness(results)
    thr = f["threshold_coverage"]
    q = f["queue_coverage"]
    ts, qs = f["threshold_summary"], f["queue_summary"]
    fig, ax = plt.subplots(figsize=(11, 8.5))
    fig.patch.set_facecolor(SURFACE)
    _style_axes(ax)

    x = np.arange(len(thr))
    w = 0.38
    ax.bar(x - w / 2, thr["catch_rate"], width=w, color=ALERT_AMBER,
           edgecolor=SURFACE, linewidth=1.5,
           label=f"Cost threshold t*={f['threshold_star']:.3f}")
    ax.bar(x + w / 2, q["catch_rate"], width=w, color=CLEARED_GREEN,
           edgecolor=SURFACE, linewidth=1.5,
           label=f"Review queue ({f['queue_capacity']} reviews)")
    ax.axhline(ts["overall_recall"], color=ALERT_AMBER, linewidth=1.0, linestyle=":")
    ax.axhline(qs["overall_recall"], color=CLEARED_GREEN, linewidth=1.0, linestyle=":")
    # fraud count per segment, placed just above the taller of the two bars so
    # it never collides with the rotated x-axis labels
    tops = np.maximum(thr["catch_rate"].to_numpy(), q["catch_rate"].to_numpy())
    for xi, nf, top in zip(x, thr["n_fraud"], tops, strict=True):
        ax.annotate(f"n={int(nf)}", (xi, top), textcoords="offset points",
                    xytext=(0, 4), ha="center", fontsize=8, color=MUTED)
    ax.set_xticks(x)
    ax.set_xticklabels(thr["segment"], rotation=35, ha="right", fontsize=9, color=INK_2)
    ax.set_ylabel("Fraud-catch coverage (share of that segment's fraud caught)")
    ax.set_ylim(0, 1)
    ax.set_title(
        "Queue fairness - per-segment fraud-catch coverage on the test window\n"
        f"Cost threshold: {ts['best_segment']} {ts['best_catch_rate']:.0%} vs "
        f"{ts['worst_segment']} {ts['worst_catch_rate']:.0%} "
        f"(a {ts['coverage_gap']:.0%} gap); dotted lines = overall recall",
        fontsize=11,
    )
    leg = ax.legend(loc="upper right", frameon=True, fontsize=10)
    leg.get_frame().set_edgecolor(BASELINE)
    note = (
        "HONEST READ-OUT: a cost-weighted policy concentrates on high-value segments by "
        "design, so\n'equal dollars' is not 'equal protection'. The queue's coverage floor "
        "guards review HEADCOUNT per\nsegment (>=5 reviews), not fraud-catch RATE - low-value "
        "segments can still be caught rarely. Labels\nare synthetic; 'caught' assumes a "
        "reviewed/flagged fraud is resolved, the same optimistic assumption\nthe cost model makes."
    )
    fig.text(0.07, 0.02, _esc(note), fontsize=9, color=INK_2)
    fig.tight_layout(rect=(0, 0.10, 1, 1))
    pdf.savefig(fig)
    plt.close(fig)


def _drift_page(pdf: PdfPages, results: dict) -> None:
    from fdo.drift import draw_drift_chart, drift_from_results

    analysis = drift_from_results(results)
    fm = analysis["fraud_mix"]
    fig, ax = plt.subplots(figsize=(11, 8.5))
    fig.patch.set_facecolor(SURFACE)
    draw_drift_chart(analysis, ax)
    note = (
        "HONEST FINDING: every input feature and the model score sit far below the 0.10 band -\n"
        "the generator's day-60 drift changes P(fraud | x), not the feature mix, and input/score\n"
        "PSI is blind to that by construction. The label-aware monitor on confirmed-fraud rows\n"
        f"is what sees it: fraud category-mix PSI {fm['category_psi']:.3f} (gift-card share "
        f"{fm['mix'].set_index('category').loc['gift_cards', 'train_share']:.0%} -> "
        f"{fm['mix'].set_index('category').loc['gift_cards', 'recent_share']:.0%}), "
        f"night share {fm['night_share_train']:.0%} -> {fm['night_share_recent']:.0%}."
    )
    fig.text(0.07, 0.02, note, fontsize=9, color=INK_2)
    fig.tight_layout(rect=(0, 0.10, 1, 1))
    pdf.savefig(fig)
    plt.close(fig)


def build_pdf(results: dict, path: str) -> str:
    with PdfPages(path) as pdf:
        _cover_page(pdf, results)
        _pr_page(pdf, results)
        _reliability_page(pdf, results)
        _threshold_page(pdf, results)
        _queue_page(pdf, results)
        _fairness_page(pdf, results)
        _drift_page(pdf, results)
    return path


# ---------------------------------------------------------------------------
# Excel workbook
# ---------------------------------------------------------------------------

_HDR = Font(bold=True, size=11)
_TITLE = Font(bold=True, size=13)


def _sheet_writer(ws):
    def put(row: int, col: int, value, font: Font | None = None, fmt: str | None = None):
        cell = ws.cell(row=row, column=col, value=value)
        if font:
            cell.font = font
        if fmt:
            cell.number_format = fmt
        cell.alignment = Alignment(vertical="top")
        return cell

    return put


def _autosize(ws, widths: dict[int, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def _metrics_sheet(wb: Workbook, results: dict) -> None:
    ws = wb.active
    ws.title = "Metrics"
    put = _sheet_writer(ws)
    m = results["metrics"]
    put(1, 1, "Model metrics (synthetic seeded data; time-based split)", _TITLE)
    put(2, 1, "PR-AUC is the primary metric: at ~1.5% prevalence accuracy is meaningless and "
              "a random scorer's PR-AUC equals the prevalence.", None)
    r = 4
    put(r, 1, "Metric", _HDR)
    put(r, 2, "Value", _HDR)
    put(r, 3, "Notes", _HDR)
    rows = [
        ("Primary model", m["primary_model"], "chosen on VALIDATION PR-AUC only"),
        ("Test PR-AUC (primary)", m["test_primary"]["pr_auc"], "held-out final 15% of the timeline"),
        ("Test PR-AUC (weighted BCE)", m["test_by_model"]["weighted_bce"]["pr_auc"], ""),
        ("Test PR-AUC (focal, gamma=2)", m["test_by_model"]["focal_gamma2"]["pr_auc"], ""),
        ("Baseline: prevalence-random PR-AUC", m["baseline_random"]["pr_auc"],
         f"theoretical floor = prevalence = {m['baseline_random']['theoretical_pr_auc']:.4f}"),
        ("Baseline: amount>p99 rule PR-AUC", m["baseline_p99"]["pr_auc"],
         f"cutoff ${m['baseline_p99']['cutoff']:.2f} fit on train; "
         f"precision {m['baseline_p99']['flag_precision']:.3f}, "
         f"recall {m['baseline_p99']['flag_recall']:.3f}"),
        ("Oracle ceiling PR-AUC", m["oracle_test_pr_auc"],
         "PR-AUC of the generator's own probabilities; knowable only because data is synthetic"),
        ("Test ROC-AUC (primary)", m["test_primary"]["roc_auc"], ""),
        ("Test precision@100", m["test_primary"].get("precision_at_100"), ""),
        ("Test precision@500", m["test_primary"].get("precision_at_500"), ""),
        ("Test prevalence", m["test_primary"]["prevalence"], ""),
        ("Val ECE before calibration", m["calibration"]["val_ece_raw"],
         "class-weighted training inflates probabilities on purpose"),
        ("Val ECE after Platt", m["calibration"]["val_ece_cal"], "calibrated on validation logits"),
        ("Test ECE before calibration", m["calibration"]["test_ece_raw"], ""),
        ("Test ECE after Platt", m["calibration"]["test_ece_cal"], ""),
        ("Val Brier before -> after", f"{m['calibration']['val_brier_raw']:.5f} -> "
                                      f"{m['calibration']['val_brier_cal']:.5f}", ""),
        ("Test Brier before -> after", f"{m['calibration']['test_brier_raw']:.5f} -> "
                                       f"{m['calibration']['test_brier_cal']:.5f}", ""),
        ("Platt (a, b)", f"({m['calibration']['platt_a']:.4f}, {m['calibration']['platt_b']:.4f})",
         "a > 0 so ranking metrics are unchanged by calibration"),
    ]
    for name, val, note in rows:
        r += 1
        put(r, 1, name)
        put(r, 2, val, fmt="0.0000" if isinstance(val, float) else None)
        put(r, 3, note)
    _autosize(ws, {1: 38, 2: 22, 3: 78})
    ws.freeze_panes = "A5"


def _thresholds_sheet(wb: Workbook, results: dict) -> None:
    ws = wb.create_sheet("Thresholds")
    put = _sheet_writer(ws)
    t = results["thresholds"]
    put(1, 1, "Cost-sensitive alert threshold", _TITLE)
    r = 2
    for line in t["costs"].describe():
        put(r, 1, line)
        r += 1
    r += 1
    summary = [
        ("Chosen threshold t* (on validation)", t["threshold_star"]),
        ("Test cost at t*", t["test_at_star"]["total_cost"]),
        ("Test cost at naive 0.5", t["test_at_naive_05"]["total_cost"]),
        ("Test cost reduction vs 0.5", t["test_cost_reduction_vs_05"]),
        ("Test cost reviewing nothing", t["test_review_none"]["total_cost"]),
        ("Test cost reviewing everything", t["test_review_all"]["total_cost"]),
        ("Test flags at t*", t["test_at_star"]["n_flagged"]),
        ("Test fraud recall at t*", t["test_at_star"]["recall"]),
    ]
    put(r, 1, "Summary", _HDR)
    for name, val in summary:
        r += 1
        put(r, 1, name)
        put(r, 2, val, fmt="#,##0.0000")
    r += 2
    put(r, 1, "Validation sweep (threshold grid)", _HDR)
    r += 1
    cols = ["threshold", "n_flagged", "review_cost_total", "missed_fraud_loss",
            "total_cost", "n_fraud_caught", "n_fraud_missed", "recall"]
    for j, c in enumerate(cols, start=1):
        put(r, j, c, _HDR)
    for _, row in t["sweep_val"].iterrows():
        r += 1
        for j, c in enumerate(cols, start=1):
            put(r, j, float(row[c]), fmt="#,##0.0000" if c == "threshold" else "#,##0.00")
    _autosize(ws, {1: 38, 2: 18, 3: 18, 4: 18, 5: 14, 6: 16, 7: 16, 8: 12})


def _queue_sheet(wb: Workbook, results: dict) -> None:
    ws = wb.create_sheet("QueuePlan")
    put = _sheet_writer(ws)
    q = results["queue"]
    df = results["df"]
    te = results["splits"]["test"]
    put(1, 1, "Analyst review-queue plan (test shift)", _TITLE)
    r = 2
    for line in q["config"].describe():
        put(r, 1, line)
        r += 1
    put(r, 1, "HONESTY NOTE: without coverage floors the LP reduces exactly to top-K by "
              f"p x amount (verified this run: {q['lp_equals_topk_ev']}).")
    r += 2
    put(r, 1, "Strategy comparison", _HDR)
    r += 1
    for j, c in enumerate(["strategy", "expected_recovered", "realized_recovered",
                           "n_reviews", "segments_covered"], start=1):
        put(r, j, c, _HDR)
    for _, row in q["comparison"].iterrows():
        r += 1
        put(r, 1, row["strategy"])
        put(r, 2, float(row["expected_recovered"]), fmt="#,##0.00")
        put(r, 3, float(row["realized_recovered"]), fmt="#,##0.00")
        put(r, 4, int(row["n_reviews"]))
        put(r, 5, int(row["segments_covered"]))
    r += 2
    put(r, 1, "Per-segment coverage", _HDR)
    r += 1
    for j, c in enumerate(q["coverage"].columns, start=1):
        put(r, j, c, _HDR)
    for _, row in q["coverage"].iterrows():
        r += 1
        for j, c in enumerate(q["coverage"].columns, start=1):
            put(r, j, row[c] if isinstance(row[c], str) else int(row[c]))
    r += 2
    put(r, 1, "Selected reviews (constrained optimizer). 'is_fraud' is shown for audit only - "
              "it is NOT known at decision time.", _HDR)
    r += 1
    header = ["rank_by_ev", "timestamp", "merchant_segment", "amount", "p_fraud_calibrated",
              "expected_value", "is_fraud"]
    for j, c in enumerate(header, start=1):
        put(r, j, c, _HDR)
    sel = q["selected_milp"]
    p_test = results["p_test"]
    sub = df.iloc[te].reset_index(drop=True)
    ev = p_test * sub["amount"].to_numpy()
    order = sel[np.argsort(-ev[sel])]
    for rank, i in enumerate(order, start=1):
        r += 1
        put(r, 1, rank)
        put(r, 2, str(sub.loc[i, "timestamp"]))
        put(r, 3, sub.loc[i, "merchant_cat"])
        put(r, 4, float(sub.loc[i, "amount"]), fmt="#,##0.00")
        put(r, 5, float(p_test[i]), fmt="0.0000")
        put(r, 6, float(ev[i]), fmt="#,##0.00")
        put(r, 7, int(sub.loc[i, "is_fraud"]))
    _autosize(ws, {1: 46, 2: 22, 3: 20, 4: 14, 5: 18, 6: 16, 7: 12})


def _fairness_sheet(wb: Workbook, results: dict) -> None:
    from fdo.fairness import queue_fairness

    ws = wb.create_sheet("Fairness")
    put = _sheet_writer(ws)
    f = queue_fairness(results)
    put(1, 1, "Queue fairness - per-segment fraud-catch coverage (test window)", _TITLE)
    put(2, 1, "Catch rate = share of that segment's OBSERVED fraud that the decision flags/reviews. "
              "Labels are synthetic; a segment with no observed fraud has a blank catch rate.")
    r = 4
    blocks = [
        (f"Cost threshold t* = {f['threshold_star']:.3f}", f["threshold_coverage"],
         f["threshold_summary"]),
        (f"Review queue ({f['queue_capacity']} reviews)", f["queue_coverage"],
         f["queue_summary"]),
    ]
    cols = ["segment", "pool_size", "n_fraud", "fraud_value", "n_fraud_caught",
            "fraud_value_caught", "catch_rate"]
    for title, cov, summ in blocks:
        put(r, 1, title, _HDR)
        r += 1
        for j, c in enumerate(cols, start=1):
            put(r, j, c, _HDR)
        for _, row in cov.iterrows():
            r += 1
            put(r, 1, row["segment"])
            put(r, 2, int(row["pool_size"]))
            put(r, 3, int(row["n_fraud"]))
            put(r, 4, float(row["fraud_value"]), fmt="#,##0.00")
            put(r, 5, int(row["n_fraud_caught"]))
            put(r, 6, float(row["fraud_value_caught"]), fmt="#,##0.00")
            cr = row["catch_rate"]
            put(r, 7, None if cr != cr else float(cr), fmt="0.0%")  # NaN -> blank
        r += 1
        put(r, 1, f"Overall recall {summ['overall_recall']:.2f}; best {summ['best_segment']} "
                  f"{summ['best_catch_rate']:.0%} vs worst {summ['worst_segment']} "
                  f"{summ['worst_catch_rate']:.0%} (gap {summ['coverage_gap']:.0%}); "
                  f"zero fraud caught in: {', '.join(summ['segments_zero_coverage']) or 'none'}.")
        r += 2
    put(r, 1, "HONEST READ-OUT: a cost-weighted policy concentrates on high-value segments by "
              "design; the coverage floor guards review headcount per segment, not fraud-catch "
              "rate. Equal dollars is not equal protection.")
    _autosize(ws, {1: 26, 2: 12, 3: 12, 4: 16, 5: 16, 6: 18, 7: 12})


def _assumptions_sheet(wb: Workbook, results: dict) -> None:
    ws = wb.create_sheet("Assumptions")
    put = _sheet_writer(ws)
    put(1, 1, "Every assumption in this workbook, in one place", _TITLE)
    t = results["thresholds"]
    q = results["queue"]
    lines = [
        "DATA: all transactions are synthetic, generated by a seeded generator "
        f"(fdo/data.py, seed={results['seed']}, n={len(results['df'])}). No real data anywhere.",
        "DATA: fraud patterns (night-time, device change, velocity, gift-card/digital-goods "
        "concentration, new accounts, high amounts) are CONSTRUCTED, so a model can learn them "
        "by design. Label noise (4% of fraud unreported, 0.08% false disputes) and mild "
        "pattern drift after day 60 are also constructed.",
        "EVALUATION: time-based split - train on the first 70% of the timeline, validate on the "
        "next 15%, test on the final 15%. The test window is used only for final reporting.",
        *t["costs"].describe(),
        *q["config"].describe(),
        "QUEUE: 'expected recovered value' = sum of p_i x amount_i over reviewed transactions, "
        "using calibrated probabilities. Realized value uses observed labels and is reported "
        "next to it as an audit.",
        "MODEL: logistic regression trained with class-weighted BCE and focal loss; the better "
        "of the two on VALIDATION PR-AUC is reported as primary.",
        "LIMITATION: no adversarial adaptation - real fraudsters change behavior when rules "
        "change; nothing here models that feedback loop.",
    ]
    r = 3
    for line in lines:
        cell = put(r, 1, line)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 30 if len(line) > 110 else 15
        r += 1
    _autosize(ws, {1: 120})


def build_excel(results: dict, path: str) -> str:
    wb = Workbook()
    _metrics_sheet(wb, results)
    _thresholds_sheet(wb, results)
    _queue_sheet(wb, results)
    _fairness_sheet(wb, results)
    _assumptions_sheet(wb, results)
    wb.save(path)
    return path


def build_deliverables(results: dict, outdir: str) -> dict[str, int]:
    """Write both deliverables and hard-verify each is above 10 KB."""
    os.makedirs(outdir, exist_ok=True)
    pdf_path = os.path.join(outdir, "executive_report.pdf")
    xlsx_path = os.path.join(outdir, "fraud_ops_workbook.xlsx")
    build_pdf(results, pdf_path)
    build_excel(results, xlsx_path)
    sizes = {p: os.path.getsize(p) for p in (pdf_path, xlsx_path)}
    for p, size in sizes.items():
        if size < MIN_BYTES:
            raise RuntimeError(f"deliverable too small ({size} bytes): {p}")
    return sizes
